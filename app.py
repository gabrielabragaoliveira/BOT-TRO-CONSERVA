import streamlit as st
import pdfplumber
import fitz  # PyMuPDF
import pandas as pd
import re
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import pytesseract

# --- 1. CONFIGURAÇÃO INTELIGENTE DO OCR ---
caminho_tesseract_win = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
if os.path.exists(caminho_tesseract_win):
    pytesseract.pytesseract.tesseract_cmd = caminho_tesseract_win

st.set_page_config(page_title="Gerador de Relatório TRO", page_icon="📸", layout="wide")
st.title("📸 Gerador de Relatório TRO")

# --- 2. PADRÕES REGEX ---
# Atualizado para ignorar espaços acidentais em volta do + ou vírgula
padrao_base = r"([A-Z]{2}-\d{3}-[A-Z]{2})\s+(.*?)\s+(Sentido\s+(?:Crescente|Decrescente))\s*-\s*([\d.,+]+)\s+([\d.,+]+)"
padrao_km_solto = r"(\d{1,4}\s*[.,+]\s*\d{1,3})"

def extrair_padrao_texto(texto_cru):
    dados = []
    if texto_cru:
        texto_limpo = re.sub(r'\s+', ' ', texto_cru)
        encontros = re.finditer(padrao_base, texto_limpo)
        for match in encontros:
            km_inicial = match.group(4)
            texto_formatado = f"{match.group(1)} {match.group(2).strip()} {match.group(3)} - {km_inicial} {match.group(5)}"
            km_chave = re.sub(r'\s+', '', km_inicial).replace(',', '.').replace('+', '.') 
            dados.append({"texto": texto_formatado, "km_chave": km_chave})
    return dados

# --- 3. MAPEAMENTO AVANÇADO DE FOTOS (COM ORDEM DE LEITURA E FILTRO) ---
def mapear_fotos(arquivos_upload):
    mapa_arquivos = {}
    
    for arquivo in arquivos_upload:
        nome_arquivo = arquivo.name
        conteudo = arquivo.read()
        extensao = nome_arquivo.split('.')[-1].lower()
        
        km_no_titulo = re.search(padrao_km_solto, nome_arquivo)
        
        # --- PROCESSAMENTO DE PDF ---
        if extensao == 'pdf':
            doc = fitz.open(stream=conteudo, filetype="pdf")
            
            for pagina in doc:
                # 1. Pega todas as imagens e filtra as que são logos pequenos
                imagens_info = pagina.get_image_info(xrefs=True)
                imagens_validas = []
                
                for img in imagens_info:
                    x0, y0, x1, y1 = img["bbox"]
                    largura, altura = x1 - x0, y1 - y0
                    # Ignora logos da ANTT, ícones e cabeçalhos (menores que 80x80 px)
                    if largura > 80 and altura > 80:
                        imagens_validas.append({"xref": img.get("xref"), "y0": y0})
                
                # Ordena as fotos reais de cima para baixo na página
                imagens_validas = sorted(imagens_validas, key=lambda x: x["y0"])

                # 2. Varre os blocos de texto buscando KMs
                blocos_texto = pagina.get_text("dict").get("blocks", [])
                kms_na_pagina = []
                
                for b in blocos_texto:
                    if b.get("type") == 0:
                        texto_bloco = ""
                        for linha in b.get("lines", []):
                            for span in linha.get("spans", []):
                                texto_bloco += span.get("text", "") + " "
                        
                        match = re.search(padrao_km_solto, texto_bloco)
                        if match:
                            km_bruto = match.group(1)
                            km_normalizado = re.sub(r'\s+', '', km_bruto).replace(',', '.').replace('+', '.')
                            y0_texto = b["bbox"][1]
                            kms_na_pagina.append({"km": km_normalizado, "y0": y0_texto})
                
                # Ordena os KMs encontrados de cima para baixo
                kms_na_pagina = sorted(kms_na_pagina, key=lambda x: x["y0"])

                # 3. Emparelha KMS e Imagens (O 1º KM da página ganha a 1ª Foto da página)
                # O min() garante que se a página tiver 3 fotos e 2 KMs, ele não quebra.
                for i in range(min(len(kms_na_pagina), len(imagens_validas))):
                    km_atual = kms_na_pagina[i]["km"]
                    xref_atual = imagens_validas[i]["xref"]
                    try:
                        imagem_bytes = doc.extract_image(xref_atual)["image"]
                        mapa_arquivos[km_atual] = imagem_bytes
                    except Exception:
                        pass

        # --- PROCESSAMENTO DE IMAGEM SOLTA (JPG/PNG) ---
        elif extensao in ['jpg', 'jpeg', 'png']:
            km_encontrado = None
            if km_no_titulo:
                km_encontrado = re.sub(r'\s+', '', km_no_titulo.group(1)).replace(',', '.').replace('+', '.')
            else:
                try:
                    img_pil = PILImage.open(io.BytesIO(conteudo))
                    texto_imagem = pytesseract.image_to_string(img_pil)
                    match_ocr = re.search(padrao_km_solto, texto_imagem)
                    if match_ocr:
                        km_encontrado = re.sub(r'\s+', '', match_ocr.group(1)).replace(',', '.').replace('+', '.')
                except Exception:
                    pass
            
            if km_encontrado:
                mapa_arquivos[km_encontrado] = conteudo
                
    return mapa_arquivos

# --- 4. INTERFACE E AÇÃO ---
texto_base = st.text_area("📝 Texto Base (Cole aqui as rodovias e KMs):", height=150)

col1, col2 = st.columns(2)
with col1:
    arquivos_antes = st.file_uploader("📸 Fotos ANTES (PDF, JPG, PNG)", accept_multiple_files=True, type=['pdf', 'jpg', 'jpeg', 'png'])
with col2:
    arquivos_depois = st.file_uploader("📸 Fotos DEPOIS (PDF, JPG, PNG)", accept_multiple_files=True, type=['pdf', 'jpg', 'jpeg', 'png'])

if st.button("🚀 Gerar Planilha Completa", use_container_width=True):
    if not texto_base.strip():
        st.warning("⚠️ Por favor, cole o texto base antes de processar.")
    else:
        with st.spinner("🔍 Lendo PDFs, ignorando logos e emparelhando fotos na ordem correta..."):
            registros = extrair_padrao_texto(texto_base)
            mapa_antes = mapear_fotos(arquivos_antes) if arquivos_antes else {}
            mapa_depois = mapear_fotos(arquivos_depois) if arquivos_depois else {}

            if registros:
                wb = Workbook()
                ws = wb.active
                ws.title = "Relatório Fotográfico"
                
                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 45 
                ws.column_dimensions['C'].width = 45 
                ws.column_dimensions['D'].width = 5
                
                ws.merge_cells('B1:C1')
                ws['B1'] = "RELATÓRIO DE EXECUÇÃO"
                ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
                ws['B1'].font = Font(bold=True, color="FFFFFF")
                ws['B1'].fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
                
                ws['B2'] = "Foto(s) do Local Antes / Durante a Execução"
                ws['C2'] = "Foto(s) do Local Após a Execução"
                for celula in ['B2', 'C2']:
                    ws[celula].alignment = Alignment(horizontal='center', vertical='center')
                    ws[celula].font = Font(bold=True)
                    ws[celula].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                
                linha_excel = 3
                fotos_coladas = 0
                
                for reg in registros:
                    texto_final = reg['texto']
                    km_buscado = reg['km_chave']
                    
                    ws.row_dimensions[linha_excel].height = 150
                    
                    foto_antes_bytes = next((img for km, img in mapa_antes.items() if km in km_buscado or km_buscado.startswith(km)), None)
                    foto_depois_bytes = next((img for km, img in mapa_depois.items() if km in km_buscado or km_buscado.startswith(km)), None)
                    
                    if foto_antes_bytes:
                        img_excel = ExcelImage(io.BytesIO(foto_antes_bytes))
                        img_excel.height, img_excel.width = 180, 300
                        ws.add_image(img_excel, f"B{linha_excel}")
                        fotos_coladas += 1
                    else:
                        ws[f"B{linha_excel}"] = "[ SEM FOTO ANTES ]"
                        ws[f"B{linha_excel}"].alignment = Alignment(horizontal='center', vertical='center')

                    if foto_depois_bytes:
                        img_excel = ExcelImage(io.BytesIO(foto_depois_bytes))
                        img_excel.height, img_excel.width = 180, 300
                        ws.add_image(img_excel, f"C{linha_excel}")
                        fotos_coladas += 1
                    else:
                        ws[f"C{linha_excel}"] = "[ SEM FOTO DEPOIS ]"
                        ws[f"C{linha_excel}"].alignment = Alignment(horizontal='center', vertical='center')

                    linha_texto_excel = linha_excel + 1
                    ws.row_dimensions[linha_texto_excel].height = 11.25
                    ws.merge_cells(start_row=linha_texto_excel, start_column=2, end_row=linha_texto_excel, end_column=3)
                    ws.cell(row=linha_texto_excel, column=2, value=texto_final).alignment = Alignment(horizontal='center', vertical='center')
                    
                    linha_excel += 2
                
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                
                st.success(f"✅ Sucesso! {len(registros)} blocos criados e {fotos_coladas} fotos inseridas na planilha.")
                st.download_button(label="📥 Baixar Relatório Final em Excel", data=excel_buffer.getvalue(), file_name="Relatorio_Antes_Depois.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            else:
                st.error("⚠️ Nenhum KM válido foi encontrado no texto base.")
